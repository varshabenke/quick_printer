from fastapi import FastAPI, UploadFile, File, Form
from fastapi.responses import HTMLResponse, JSONResponse
from sqlalchemy import create_engine, Column, String, Integer
from sqlalchemy.orm import sessionmaker, declarative_base
import razorpay
import uuid
import os
import win32api
import win32print
import time
from PIL import Image
from fastapi import Query
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from fastapi import UploadFile, File

app = FastAPI()
os.makedirs("uploads", exist_ok=True)

# 🔑 Razorpay keys
RAZORPAY_KEY_ID = "rzp_test_Snyahca9JBENJE"
RAZORPAY_KEY_SECRET = "XhyT5QIixQsEtrFtbMQIJKNT"

client = razorpay.Client(auth=(RAZORPAY_KEY_ID, RAZORPAY_KEY_SECRET))
latest_uploaded_file = ""
latest_uploaded_filename = ""

# ---------------- DB ----------------
DATABASE_URL = "sqlite:///./orders.db"
engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(bind=engine)
Base = declarative_base()

class Order(Base):
    __tablename__ = "orders"

    id = Column(String, primary_key=True)
    file_path = Column(String)
    copies = Column(Integer)
    pages = Column(Integer)
    amount = Column(Integer)
    status = Column(String)
    razorpay_order_id = Column(String)

Base.metadata.create_all(bind=engine)

UPLOAD_FOLDER = "uploads"


# ---------------- HOME ----------------
@app.get("/", response_class=HTMLResponse)
def home():
    with open("index.html", "r", encoding="utf-8") as f:
        return f.read()

# ---------------- UPLOAD + CREATE ORDER ----------------


@app.post("/upload")
async def upload_file(
    file: UploadFile = File(...),
    pages: int = Form(1),
    copies: int = Form(1),
    print_type: str = Form("bw"),
    layout: str = Form("portrait")
):

    global latest_uploaded_file
    global latest_uploaded_filename
   

    try:

        print("UPLOAD HIT")

        file_id = str(uuid.uuid4())

        file_path = os.path.join(
            "uploads",
            f"{file_id}_{file.filename}"
        )

        content = await file.read()

        with open(file_path, "wb") as f:
            f.write(content)

        latest_uploaded_file = file_path
        latest_uploaded_filename = file.filename

        print("Saved:", latest_uploaded_file)

        price = 2 if print_type == "bw" else 5

        amount = pages * copies * price

        razorpay_order = client.order.create({
            "amount": amount * 100,
            "currency": "INR",
            "payment_capture": 1
        })
        
        new_order = Order(

            id=file_id,

            file_path=file_path,

            copies=copies,

            pages=pages,

            amount=amount,

            status="pending",

            razorpay_order_id=razorpay_order["id"]

        )
        db = SessionLocal()
        db.add(new_order)

        db.commit()

        print("ORDER SAVED:", file_id)

        return {
            "order_id": file_id,
            "razorpay_order_id": razorpay_order["id"],
            "amount": amount,
            "key": RAZORPAY_KEY_ID
        }

    except Exception as e:

        print("UPLOAD ERROR:", e)

        return {
            "error": str(e)
        }
# ---------------- VERIFY PAYMENT (FRONTEND CONFIRMATION) ----------------
@app.post("/verify-payment")
async def verify_payment(data: dict):

    payment_id = data.get("payment_id")
    order_id = data.get("order_id")

    try:
        payment = client.payment.fetch(payment_id)

        if payment["status"] == "captured":

            db = SessionLocal()
            order = db.query(Order).filter(Order.id == order_id).first()

            if order:
                order.status = "paid"
                db.commit()

                # AUTO PRINT
                try:
                    os.startfile(order.file_path, "print")
                except Exception as e:
                    print("PRINT ERROR:", e)

            db.close()

            return {"status": "success"}

        return {"status": "failed"}

    except Exception as e:
        return {"error": str(e)}


def print_file(file_path, copies, print_type):

    ext = os.path.splitext(file_path)[1].lower()

    try:

        # IMAGE FILES
        if ext in [".jpg", ".jpeg", ".png"]:

            image = Image.open(file_path)

            pdf_path = file_path + ".pdf"

            # BLACK & WHITE
            if print_type == "bw":

                image = image.convert("L")
            
            else:

                image = image.convert("RGB")

            image.save(pdf_path, "PDF")

            file_path = pdf_path

        # PRINT COPIES
        printer_name = win32print.GetDefaultPrinter()
        for i in range(copies):

            try:

                win32api.ShellExecute(
                    0,
                    "printto",
                    file_path,
                    f'"{printer_name}"',
                    ".",
                    0
                )
               
                print("PRINT SENT")

            except Exception as e:

                print("PRINT ERROR:", e)

            time.sleep(3)
        return "print sent"
    except Exception as e:

        print("CRITICAL PRINT ERROR:", e)

        return "Failed"

@app.get("/success/{order_id}", response_class=HTMLResponse)
def success(
    order_id: str,
    payment_id: str = Query(""),
    copies: int = Query(1),
    pages: int = Query(1),
    layout: str = Query("portrait"),
    amount: int = Query(0),
    print_type: str = Query("bw")
):

    global latest_uploaded_file
    global latest_uploaded_filename
    print("SUCCESS ROUTE HIT")

    db = SessionLocal()

    order = db.query(Order).filter(
        Order.id == order_id
    ).first()

    if not order:

        return HTMLResponse("""
        <h2>

            "error": "Order not found"
        </h2>

        """)



    if order.status == "printed":

        return HTMLResponse("""

        <h2>
        ✅ Already Printed
        </h2>

        """)
    # START PRINTING
    print("AUTO PRINT FILE:", latest_uploaded_file)

    print_file(latest_uploaded_file,copies,print_type)

    # UPDATE STATUS
    order.status = "printed"

    db.commit()

    # IMPORTS
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from datetime import datetime
    import os

    # CURRENT TIME
    current_time = datetime.now().strftime(
        "%Y-%m-%d %H:%M:%S"
    )

    # EXCEL FILE
    file_name = "transactions.xlsx"

    # HEADERS
    headers = [
        "File Name",
        "Token No",
        "Transaction ID",
        "Date & Time",
        "Copies",
        "Pages",
        "Layout",
        "Amount",
        "Print Status"
    ]
    if os.path.exists(file_name):

        wb_temp = load_workbook(file_name)

        ws_temp = wb_temp.active

        token_number = ws_temp.max_row

    else:

        token_number = 1

    try:

        print_file(latest_uploaded_file,copies,print_type)

        print_status = "Printed"

    except Exception as e:

        print_status = "Failed"

        print("Print Error:", e)

    # DATA
    data_row = [
        latest_uploaded_filename,
        token_number,
        payment_id,
        current_time,
        copies,
        pages,
        layout,
        amount,
        print_status
    ]

    # BORDER STYLE
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # CREATE NEW FILE
    if not os.path.exists(file_name):

        wb = Workbook()

        ws = wb.active

        ws.title = "Transactions"

        # ADD HEADERS
        ws.append(headers)

        # STYLE HEADERS
        for cell in ws[1]:

            cell.font = Font(bold=True)

            cell.alignment = Alignment(
                horizontal="center"
            )

            cell.border = thin_border

        # ADD FIRST DATA ROW
        ws.append(data_row)

        # STYLE DATA ROW
        for row in ws.iter_rows(
            min_row=2,
            max_row=2
        ):

            for cell in row:

                cell.alignment = Alignment(
                    horizontal="center"
                )

                cell.border = thin_border

        # COLUMN WIDTHS
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 30

        wb.save(file_name)

    # APPEND EXISTING FILE
    else:

        wb = load_workbook(file_name)

        ws = wb.active

        ws.append(data_row)

        last_row = ws.max_row

        # STYLE LAST ROW
        for row in ws.iter_rows(
            min_row=last_row,
            max_row=last_row
        ):

            for cell in row:

                cell.alignment = Alignment(
                    horizontal="center"
                )

                cell.border = thin_border

        wb.save(file_name)

    # SUCCESS PAGE
    return f"""

    <html>

    <head>

        <title>Printing Started</title>

        <style>

            body{{
                font-family:Arial;
                text-align:center;
                padding-top:100px;
                background:#f5f5f5;
            }}

            .box{{
                background:white;
                width:350px;
                margin:auto;
                padding:40px;
                border-radius:15px;
                box-shadow:0 0 10px rgba(0,0,0,0.1);
            }}

            h1{{
                color:green;
            }}

        </style>

    </head>

    <body>

        <div class="box">

            <h1>✅ Payment Successful</h1>

            <p>
                <b>Transaction ID:</b><br>
                {payment_id}
            </p>

            <p>
                <b>Copies:</b> {copies}
            </p>

            <p>
                <b>Pages:</b> {pages}
            </p>

            <p>
                <b>Layout:</b> {layout}
            </p>

            <h3>🖨 Printing Started...</h3>

        </div>

    </body>

    </html>

    """
